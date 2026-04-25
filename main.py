import pandas as pd
from faker import Faker
import random
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

fake = Faker()
random.seed(42)

# ── 1. Generate messy data ──────────────────────────────────────────
def generate_messy_data(n=120):
    statuses = ["active", "Active", "ACTIVE", "inactive", "Inactive", "pending", "PENDING"]
    rows = []
    for _ in range(n):
        rows.append({
            "customer_name": random.choice([
                fake.name(),
                fake.name().upper(),
                "  " + fake.name() + "  ",   # extra spaces
                "",                            # blank name
            ]),
            "email": random.choice([
                fake.email(),
                fake.email().upper(),
                "not-an-email",
                "",
            ]),
            "purchase_amount": random.choice([
                round(random.uniform(10, 2000), 2),
                None,
                -round(random.uniform(1, 50), 2),   # negative = data error
            ]),
            "purchase_date": random.choice([
                fake.date_between(start_date="-1y", end_date="today").strftime("%Y-%m-%d"),
                fake.date_between(start_date="-1y", end_date="today").strftime("%d/%m/%Y"),
                "not-a-date",
                "",
            ]),
            "status": random.choice(statuses),
            "country": random.choice([
                fake.country(),
                fake.country().lower(),
                fake.country().upper(),
                "",
            ]),
        })
    # Add deliberate duplicates
    rows += random.choices(rows[:20], k=15)
    random.shuffle(rows)
    return pd.DataFrame(rows)

# ── 2. Clean the data ───────────────────────────────────────────────
def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    original_count = len(df)

    # Standardise text columns
    for col in ["customer_name", "country"]:
        df[col] = df[col].astype(str).str.strip().str.title()
        df[col] = df[col].replace({"": None, "Nan": None})

    # Standardise email
    df["email"] = df["email"].astype(str).str.strip().str.lower()
    df.loc[~df["email"].str.contains(r"@.*\.", na=False), "email"] = None

    # Standardise status
    df["status"] = df["status"].astype(str).str.strip().str.capitalize()

    # Fix purchase_amount: keep only positive numbers
    df["purchase_amount"] = pd.to_numeric(df["purchase_amount"], errors="coerce")
    df.loc[df["purchase_amount"] < 0, "purchase_amount"] = None

    # Standardise purchase_date → YYYY-MM-DD
    def parse_date(val):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return pd.to_datetime(val, format=fmt).date()
            except Exception:
                pass
        return None

    df["purchase_date"] = df["purchase_date"].apply(parse_date)

    # Drop rows missing critical fields
    df = df.dropna(subset=["customer_name", "purchase_amount"])

    # Remove duplicates
    df = df.drop_duplicates()

    removed = original_count - len(df)
    print(f"  Rows removed (duplicates / invalid): {removed}")
    print(f"  Clean rows remaining:                {len(df)}")
    return df.reset_index(drop=True)

# ── 3. Build styled Excel report ────────────────────────────────────
def build_excel_report(df: pd.DataFrame, path: str):
    summary = pd.DataFrame({
        "Metric": [
            "Total Customers",
            "Total Revenue (€)",
            "Average Order (€)",
            "Highest Order (€)",
            "Active Customers",
            "Inactive Customers",
            "Pending Customers",
            "Report Generated",
        ],
        "Value": [
            len(df),
            f"{df['purchase_amount'].sum():,.2f}",
            f"{df['purchase_amount'].mean():,.2f}",
            f"{df['purchase_amount'].max():,.2f}",
            len(df[df['status'] == 'Active']),
            len(df[df['status'] == 'Inactive']),
            len(df[df['status'] == 'Pending']),
            datetime.now().strftime("%Y-%m-%d %H:%M"),
        ]
    })

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        df.to_excel(writer, sheet_name="Clean Data", index=False)

    wb = load_workbook(path)

    # ── Style: Summary sheet ──
    ws = wb["Summary"]
    header_fill = PatternFill("solid", fgColor="2E86AB")
    alt_fill    = PatternFill("solid", fgColor="D6EAF8")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin        = Side(style="thin", color="AAAAAA")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border    = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif cell.row % 2 == 0:
                cell.fill = alt_fill

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    ws["A1"].value = "📊 Report Summary"
    ws.merge_cells("A1:B1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # ── Style: Clean Data sheet ──
    ws2 = wb["Clean Data"]
    col_widths = [22, 30, 18, 16, 14, 20]
    for i, width in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    for row in ws2.iter_rows():
        for cell in row:
            cell.border    = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F0F8FF")

    # Colour-code status column
    status_colors = {"Active": "D5F5E3", "Inactive": "FADBD8", "Pending": "FEF9E7"}
    status_col = [c.column for c in ws2[1] if c.value == "status"]
    if status_col:
        sc = status_col[0]
        for row in ws2.iter_rows(min_row=2):
            cell  = row[sc - 1]
            color = status_colors.get(str(cell.value), "FFFFFF")
            cell.fill = PatternFill("solid", fgColor=color)

    wb.save(path)
    print(f"\n  ✅ Report saved → {path}")

# ── 4. Run everything ───────────────────────────────────────────────
if __name__ == "__main__":
    print("\n🔄 Generating messy data...")
    messy = generate_messy_data(120)
    messy.to_csv("messy_data.csv", index=False)
    print(f"  Messy rows created: {len(messy)}")

    print("\n🧹 Cleaning data...")
    clean = clean_data(messy)

    print("\n📊 Building Excel report...")
    build_excel_report(clean, "clean_report.xlsx")

    print("\n🎉 Done! Open clean_report.xlsx to see the result.")
